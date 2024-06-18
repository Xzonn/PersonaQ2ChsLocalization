import os
from openpyxl import Workbook
from openpyxl.formatting import Rule
from openpyxl.styles import DEFAULT_FONT, Font, Alignment, PatternFill, Protection
from openpyxl.styles.differential import DifferentialStyle

from helper import DIR_CSV_ROOT, DIR_XLSX_ROOT, load_csv

LANGUAGE = os.getenv("XZ_LANGUAGE") or "zh_Hans"


def convert_csv_to_xlsx(csv_root_without_language: str, language: str, xlsx_root: str):
  _font = Font(name="微软雅黑", sz=10)
  {k: setattr(DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}

  bold = Font(name="微软雅黑", sz=10, b=True)

  center = Alignment(horizontal="center", vertical="center")
  center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
  top_left_wrap = Alignment(vertical="top", wrap_text=True)
  normal_japanese = Font(name="Meiryo", sz=10)

  gray_chinese = Font(name="微软雅黑", sz=10, color="666666")

  rule_control = Rule(type="expression", dxf=DifferentialStyle(fill=PatternFill(bgColor="ffcccc")), stopIfTrue=True)
  rule_control.formula = ["$H2"]

  protection_none = Protection(locked=False)

  for root, dirs, files in os.walk(f"{csv_root_without_language}/ja"):
    for file_name in files:
      if not file_name.endswith(".csv"):
        continue

      sheet_name = os.path.relpath(
        f"{root}/{file_name}",
        f"{csv_root_without_language}/ja",
      ).replace("\\", "/").removesuffix(".csv")
      print(f"Converting: {sheet_name}.csv")

      original = load_csv(f"{csv_root_without_language}/ja", sheet_name)
      translated = load_csv(f"{csv_root_without_language}/{language}", sheet_name)

      workbook = Workbook()
      sheet = workbook.active
      sheet.protection.sheet = True
      sheet.append((
        "id",
        "source",
        "target",
        "comment_1",
        "comment_2",
        "comment_3",
        sheet_name,
        "control_check",
      ))

      for i, (original_line, translated_line) in enumerate(zip(original, translated)):
        mt_translations = translated_line["developer_comments"]
        comments = ""
        if "\n\n评论：\n\n" in mt_translations:
          mt_translations, comments = mt_translations.split("\n\n评论：\n\n", 1)
        sheet.append((
          original_line["id"],
          original_line["target"],
          translated_line["target"],
          original_line["developer_comments"],
          mt_translations,
          comments,
          "",
          f'=OR(LEN(B{i + 2})-LEN(SUBSTITUTE(SUBSTITUTE(B{i + 2},"[",""),"]",""))<>LEN(C{i + 2})-LEN(SUBSTITUTE(SUBSTITUTE(C{i + 2},"[",""),"]","")),LEN(C{i + 2})-LEN(SUBSTITUTE(C{i + 2},"[",""))<>LEN(C{i + 2})-LEN(SUBSTITUTE(C{i + 2},"]","")))',
        ))

        for col in "ABCDEF":
          cell = sheet[f"{col}{i + 2}"]
          if col in "B":
            cell.font = normal_japanese
          elif col in "ADE":
            cell.font = gray_chinese
          if col in "CF":
            cell.protection = protection_none
          if col in "A":
            cell.alignment = center
          elif col in "D":
            cell.alignment = center_wrap
          else:
            cell.alignment = top_left_wrap

      for col in "ABCDEFGH":
        column = sheet.column_dimensions[col]
        if col in "AD":
          column.width = 15
          column.alignment = center_wrap
        elif col in "BCE":
          column.width = 32
          column.alignment = top_left_wrap
        elif col in "F":
          column.width = 48
          column.alignment = top_left_wrap
        elif col in "GH":
          column.hidden = True

      sheet.freeze_panes = "B2"
      sheet.conditional_formatting.add(f"C2:C{sheet.max_row}", rule_control)

      output_path = f"{xlsx_root}/{sheet_name}.xlsx"
      os.makedirs(os.path.dirname(output_path), exist_ok=True)
      workbook.save(output_path)


if __name__ == "__main__":
  convert_csv_to_xlsx(DIR_CSV_ROOT, LANGUAGE, f"{DIR_XLSX_ROOT}/{LANGUAGE}")
